# ====================MID_RANGE_ORG====================
"""Mid-Range Organisations — New Zealand Public Data.

Seven agencies, one platform, one app. Built from `design/app_design.md`; every
element here traces to a row of `design/mart_contract.md`.

====================SCREEN_LAYOUT====================

    ┌──────────────────────┬────────────────────────────────────────────────┐
    │ SIDEBAR              │  Mid-Range Organisations — NZ Public Data      │
    │                      │  ┌──────────────────────────────────────────┐  │
    │ 🏛️ Mid-Range Org     │  │▨▨▨▨ 45° hazard stripe, #FFD100 on #111 ▨▨│  │
    │ ──────────────       │  ├──────────────────────────────────────────┤  │
    │ Organisation         │  │ ⚠ BUILT FROM NEW ZEALAND GOVERNMENT DATA │  │
    │ ☑ DIA ☑ CUS ☑ SNZ    │  │   — NOT AN OFFICIAL GOVERNMENT PRODUCT   │  │
    │ ☑ MFT ☑ LNZ ☑ MFE    │  │   the seven agencies, named; produced    │  │
    │ ☑ MOH                │  │   independently; every source file is    │  │
    │                      │  │   listed in 🔎 Data & Provenance         │  │
    │ ──────────────       │  ├──────────────────────────────────────────┤  │
    │ PROVENANCE           │  │▨▨▨▨ 45° hazard stripe, #FFD100 on #111 ▨▨│  │
    │ ● REAL — measured    │  └──────────────────────────────────────────┘  │
    │ ● DERIVED — computed │                                                │
    │ ● COMPARE — 2nd org  │  ┌────┬────┬────┬────┬────┬────┬────┬────┐    │
    │ 🔶 SYNTHETIC         │  │ 🏛️ │ 💹 │ 🌏 │ 🎲 │ 🩺 │ 🌡️ │ 🗺️ │ 🔎 │    │
    │ ◍ SURVEY — has CIs   │  └────┴────┴────┴────┴────┴────┴────┴────┘    │
    │ ──────────────       │                                                │
    │ N source files       │   ── tab body ──                               │
    │ 7 agencies · 6 hosts │                                                │
    └──────────────────────┴────────────────────────────────────────────────┘

====================TAB_LIST====================

    1 🏛️  Overview            real            what the seven agencies publish
    2 💹  Economy             real            GDP, CPI, labour, migration
    3 🌏  Trade & Treaties    real            trade by partner; treaty register
    4 🎲  Civic & Charitable  part synthetic  gaming proceeds; charities register
    5 🩺  Health              real (survey)   NZHS indicators with CIs
    6 🌡️  Environment         real            GHG inventory; MfE river water
                                          quality; air trends; accounts
    7 🗺️  Places & Property   real            LINZ cadastre + address density;
                                          Gazetteer map; property transfers
    8 🔎  Data & Provenance   real            source register, validation, gaps
    9 🏗️  Build Notes         real            the build write-up, loaded at run
                                          time from mid_range_org__readme.md

====================VISUAL_VOCABULARY====================

    REAL       #2a78d6 blue      a figure an agency published
    DERIVED    #eb6834 orange    computed here from published figures
    COMPARE    #1baf7a aqua      a second agency's measure of the same thing
    SYNTHETIC  #e87ba4 magenta   modelled, not measured

    🔶  prefixes every synthetic figure — metric label, chart title, legend
        entry, and detail-table column header. Not the methodology tab only.
    (derived)  appended to a derived measure, method stated in the caption.
    ◍   a survey estimate; always drawn with its confidence-interval band,
        never as a bare line, never presented as a count.
    Suppressed values are a GAP, never zero; the tooltip names the symbol the
    agency published.

    Banners: a hazard-striped black-and-#FFD100 header above every tab, drawn as
    raw HTML so the stripes survive both themes and matching the MSD platform's
    banner so the two read as one family. Inside a tab, st.warning names which
    elements are modelled; st.error would mark a fully synthetic tab (there is
    none in this build).

    One y-axis, always. Two measures of different scale become two charts or an
    indexed series — never a second axis.

====================DEPARTURES_FROM_THE_TEMPLATE====================

This runs on Streamlit Community Cloud, not Streamlit in Snowflake, so
`get_active_session()` is replaced by a read-only DuckDB connection behind the
same function boundary. `df_db_schema` stays the first argument of every data
method even though DuckDB does not need it: keeping it means moving this app into
Snowflake is swapping `run_query` for `session.sql(...).to_pandas()` and nothing
else.

`get_connection` takes an extract fingerprint (path, size, mtime) as an argument.
Community Cloud hot-reloads on a push — it pulls the new files and re-runs the
script, but does **not** clear `cache_resource`. Without the fingerprint a
connection opened before the pull keeps reading the replaced file, and a query
written against a newly added column fails to bind against data sitting correctly
on disk a few bytes away. Making it an argument turns a data refresh into a cache
miss.
"""

# ====================IMPORTS====================
from __future__ import annotations

import io
import json
import os
from pathlib import Path

import duckdb
import pandas as pd
import plotly.graph_objects as go
import pydeck as pdk
import streamlit as st

st.set_page_config(
    layout="wide",
    page_title="Mid-Range Organisations — NZ Public Data",
    page_icon="🏛️",
)

# ====================PALETTE====================
# The validated dataviz reference palette. Colour carries meaning here and
# nothing else does.
REAL = "#2a78d6"
DERIVED = "#eb6834"
COMPARE = "#1baf7a"
SYNTHETIC = "#e87ba4"
SURVEY_BAND = "rgba(42,120,214,0.16)"
INK = "#10161d"
MUTED = "#74818e"
GRID = "#e5e9ef"

# Categorical hues in fixed order, never cycled. A ninth series folds to "Other".
SERIES_HUES = [REAL, DERIVED, COMPARE, "#eda100", SYNTHETIC,
               "#008300", "#4a3aa7", "#e34948"]

ORG_NAMES = {
    "DIA": "Internal Affairs", "CUS": "Customs", "SNZ": "Stats NZ",
    "MFT": "Foreign Affairs & Trade", "LNZ": "Land Information",
    "MFE": "Environment", "MOH": "Health",
}

# ====================SESSION====================
APP_DIR = Path(__file__).resolve().parent

# The Method tab renders a markdown document loaded from disk at run time, not
# embedded in this module. One copy, always current, and never a second that can
# fall out of date with the build it describes.
REFERENCE_DIRS = [
    os.environ.get("MID_RANGE_ORG_REFERENCE_DIR", ""),
    str(APP_DIR.parent / "reference"),
    str(APP_DIR.parent),
]
# Named for the app module it documents, so the pair stay obviously together.
BUILD_NOTES_DOC = "mid_range_org__readme.md"

DB_CANDIDATES = [
    APP_DIR.parent / "data" / "mid_range_org_public.duckdb",     # public_repo
    APP_DIR.parent / "public" / "mid_range_org_public.duckdb",   # working project
    Path("data/mid_range_org_public.duckdb"),                    # cwd
]


def _extract_fingerprint():
    """Path, size and mtime of the extract, used as a cache key.

    Not decoration: Community Cloud re-runs the script on a push without
    clearing `cache_resource`, so a connection opened before the pull keeps
    reading the replaced file. Passing this in makes a data refresh a cache miss.
    """
    for path in DB_CANDIDATES:
        if path.exists():
            s = path.stat()
            return (str(path), s.st_size, int(s.st_mtime))
    return None


@st.cache_resource
def _open_connection(fingerprint):
    if fingerprint is None:
        return None
    return duckdb.connect(fingerprint[0], read_only=True)


def get_connection():
    return _open_connection(_extract_fingerprint())


@st.cache_data(show_spinner=False)
def run_query(fingerprint, sql: str) -> pd.DataFrame:
    """The one place SQL meets the database.

    Moving this app into Snowflake is replacing this function body with
    `session.sql(sql).to_pandas()`.
    """
    con = _open_connection(fingerprint)
    if con is None:
        return pd.DataFrame()
    try:
        return con.execute(sql).df()
    except Exception as exc:  # noqa: BLE001
        st.error(f"Query failed: {exc}")
        return pd.DataFrame()


def q(sql: str) -> pd.DataFrame:
    return run_query(_extract_fingerprint(), sql)


def sql_list(values) -> str:
    """A SQL IN-list from a python iterable, quoted and escaped."""
    vals = [str(v).replace("'", "''") for v in values]
    return "'" + "','".join(vals) + "'" if vals else "''"


# ====================DATA====================
# Every method takes df_db_schema first, even though DuckDB ignores it.

@st.cache_data(show_spinner=False)
def get_platform_summary(df_db_schema, fingerprint):
    return run_query(fingerprint, "SELECT * FROM M_PLATFORM_SUMMARY")


@st.cache_data(show_spinner=False)
def get_org_register(df_db_schema, fingerprint, orgs):
    return run_query(fingerprint, f"""
        SELECT * FROM M_ORG_REGISTER WHERE ORG IN ({sql_list(orgs)})
        ORDER BY STAGED_ROWS DESC
    """)


@st.cache_data(show_spinner=False)
def get_series_catalog(df_db_schema, fingerprint, orgs):
    return run_query(fingerprint, f"""
        SELECT * FROM M_ECON_SERIES_CATALOG
        WHERE ORG IN ({sql_list(orgs)}) AND OBS_WITH_VALUE > 8
        ORDER BY SUBJECT, GROUP_NAME, SERIES_NAME
    """)


@st.cache_data(show_spinner=False)
def get_series(df_db_schema, fingerprint, refs, period_from, period_to):
    if not refs:
        return pd.DataFrame()
    return run_query(fingerprint, f"""
        SELECT SERIES_REF, SERIES_NAME, SUBJECT, UNITS, PERIOD, PERIOD_TYPE,
               VALUE, IS_SUPPRESSED, SUPPRESSION_SYMBOL
        FROM M_ECON_SERIES
        WHERE SERIES_REF IN ({sql_list(refs)})
          AND PERIOD >= '{period_from}' AND PERIOD <= '{period_to}'
        ORDER BY SERIES_REF, PERIOD
    """)


@st.cache_data(show_spinner=False)
def get_trade_monthly(df_db_schema, fingerprint):
    return run_query(fingerprint,
                     "SELECT * FROM M_TRADE_MONTHLY ORDER BY PERIOD, ACCOUNT")


@st.cache_data(show_spinner=False)
def get_trade_by_country(df_db_schema, fingerprint, account, period_from):
    return run_query(fingerprint, f"""
        SELECT COUNTRY_LABEL, sum(VALUE_NZD) AS VALUE_NZD
        FROM M_TRADE_BY_COUNTRY
        WHERE ACCOUNT = '{account}' AND PERIOD >= '{period_from}'
        GROUP BY ALL ORDER BY VALUE_NZD DESC
    """)


@st.cache_data(show_spinner=False)
def get_trade_by_hs2(df_db_schema, fingerprint, account, period_from):
    return run_query(fingerprint, f"""
        SELECT HS2_LABEL, sum(VALUE_NZD) AS VALUE_NZD
        FROM M_TRADE_BY_HS2
        WHERE ACCOUNT = '{account}' AND PERIOD >= '{period_from}'
        GROUP BY ALL ORDER BY VALUE_NZD DESC
    """)


@st.cache_data(show_spinner=False)
def get_treaties(df_db_schema, fingerprint, status, ttype):
    where = ["TRUE"]
    if status != "All":
        where.append(f"NZ_STATUS = '{status}'")
    if ttype != "All":
        where.append(f"TREATY_TYPE = '{ttype}'")
    return run_query(fingerprint, f"""
        SELECT TREATY_ID, TITLE, TREATY_TYPE, DATE_CONCLUDED, YEAR_CONCLUDED,
               NZ_STATUS, OTHER_PARTY
        FROM M_TREATY WHERE {' AND '.join(where)}
        ORDER BY YEAR_CONCLUDED DESC NULLS LAST
    """)


@st.cache_data(show_spinner=False)
def get_treaty_decades(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT DECADE_CONCLUDED, TREATY_TYPE, sum(TREATY_COUNT) AS TREATY_COUNT
        FROM M_TREATY_BY_DECADE GROUP BY ALL ORDER BY DECADE_CONCLUDED
    """)


@st.cache_data(show_spinner=False)
def get_gmp_national(df_db_schema, fingerprint):
    return run_query(fingerprint,
                     "SELECT * FROM M_GMP_NATIONAL ORDER BY PERIOD")


@st.cache_data(show_spinner=False)
def get_gmp_ta(df_db_schema, fingerprint, tas):
    where = f"AND TA_NAME IN ({sql_list(tas)})" if tas else ""
    return run_query(fingerprint, f"""
        SELECT PERIOD, TA_NAME, IS_CLUSTER, GMP_NZD, SHARE_OF_TOTAL,
               IS_SUPPRESSED, SUPPRESSION_SYMBOL, PUBLICATIONS,
               REVISION_NZD, WAS_REVISED
        FROM M_GMP_TA WHERE TRUE {where} ORDER BY PERIOD, TA_NAME
    """)


@st.cache_data(show_spinner=False)
def get_gmp_per_capita(df_db_schema, fingerprint, year):
    return run_query(fingerprint, f"""
        SELECT TA_NAME, GMP_NZD, POPULATION, GMP_PER_CAPITA, IS_DERIVED,
               IS_COMPLETE_YEAR
        FROM M_GMP_PER_CAPITA
        WHERE PERIOD_YEAR = '{year}' AND GMP_PER_CAPITA IS NOT NULL
          AND IS_COMPLETE_YEAR
        ORDER BY GMP_PER_CAPITA DESC
    """)


@st.cache_data(show_spinner=False)
def get_per_capita_years(df_db_schema, fingerprint):
    """Years the derived per-capita measure can actually answer."""
    return run_query(fingerprint, """
        SELECT PERIOD_YEAR, count(*) AS TA_COUNT
        FROM M_GMP_PER_CAPITA
        WHERE GMP_PER_CAPITA IS NOT NULL AND IS_COMPLETE_YEAR
        GROUP BY ALL HAVING count(*) > 10 ORDER BY PERIOD_YEAR
    """)


@st.cache_data(show_spinner=False)
def get_synthetic_venues(df_db_schema, fingerprint, period, tas):
    where = f"AND TA_NAME IN ({sql_list(tas)})" if tas else ""
    return run_query(fingerprint, f"""
        SELECT VENUE_ID, TA_NAME, VENUE_TYPE, PERIOD, MACHINES, GMP_NZD,
               GMP_PER_MACHINE, IS_GRANDFATHERED, IS_SYNTHETIC
        FROM SYN_VENUE_GMP WHERE PERIOD = '{period}' {where}
        ORDER BY GMP_NZD DESC
    """)


@st.cache_data(show_spinner=False)
def get_charity_summary(df_db_schema, fingerprint):
    return run_query(fingerprint, "SELECT * FROM M_CHARITY_SUMMARY")


@st.cache_data(show_spinner=False)
def get_charity_by_sector(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT * FROM M_CHARITY_BY_SECTOR ORDER BY CHARITY_COUNT DESC
    """)


@st.cache_data(show_spinner=False)
def get_charities(df_db_schema, fingerprint, sector, status, text):
    where = ["TRUE"]
    if sector != "All":
        where.append(f"SECTOR_NAME = '{sector}'")
    if status != "All":
        where.append(f"STATUS = '{status}'")
    if text:
        safe = text.replace("'", "''")
        where.append(f"NAME ILIKE '%{safe}%'")
    return run_query(fingerprint, f"""
        SELECT REGISTRATION_NUMBER, NAME, SECTOR_NAME, ACTIVITY_NAME,
               BENEFICIARY_NAME, STATUS, DATE_REGISTERED, CITY
        FROM M_CHARITY WHERE {' AND '.join(where)}
        ORDER BY NAME LIMIT 8000
    """)


@st.cache_data(show_spinner=False)
def get_health_topics(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT * FROM M_HEALTH_TOPIC ORDER BY TOPIC, INDICATOR
    """)


@st.cache_data(show_spinner=False)
def get_health_series(df_db_schema, fingerprint, indicator, population,
                      breakdown, age_std, geography="National"):
    """One indicator over time, split by a demographic OR by region.

    `geography="Regional council"` switches the series from demographic groups
    to regions; the extract carries region x total and region x ethnicity, which
    is exactly what this offers.
    """
    if geography == "National":
        geo_clause = "GEOGRAPHY_VALUE = 'All'"
        split = f"DEMOGRAPHIC_TYPE = '{breakdown}'"
    else:
        geo_clause = f"GEOGRAPHY_TYPE = '{geography}'"
        split = f"DEMOGRAPHIC_TYPE = '{breakdown}'"
    return run_query(fingerprint, f"""
        SELECT YEAR_FROM, YEAR_TO, DEMOGRAPHIC_TYPE, DEMOGRAPHIC_VALUE,
               GEOGRAPHY_TYPE, GEOGRAPHY_VALUE, VALUE_TYPE, VALUE, LOW, HIGH
        FROM M_HEALTH_INDICATOR
        WHERE INDICATOR = '{indicator}' AND POPULATION = '{population}'
          AND {split}
          AND lower(CAST(AGE_STANDARDISED AS VARCHAR)) = '{str(age_std).lower()}'
          AND {geo_clause} AND VALUE IS NOT NULL
        ORDER BY YEAR_TO, DEMOGRAPHIC_VALUE
    """)


@st.cache_data(show_spinner=False)
def get_health_breakdowns(df_db_schema, fingerprint, geography):
    """The breakdowns the extract can actually answer at this geography."""
    where = ("GEOGRAPHY_VALUE = 'All'" if geography == "National"
             else f"GEOGRAPHY_TYPE = '{geography}'")
    return run_query(fingerprint, f"""
        SELECT DEMOGRAPHIC_TYPE, count(*) AS OBS
        FROM M_HEALTH_INDICATOR WHERE {where}
        GROUP BY ALL ORDER BY OBS DESC
    """)


@st.cache_data(show_spinner=False)
def get_life_expectancy(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT PERIOD, AREA, SEX, LIFE_EXPECTANCY
        FROM M_LIFE_EXPECTANCY
        WHERE LIFE_EXPECTANCY IS NOT NULL
        ORDER BY PERIOD DESC, LIFE_EXPECTANCY DESC
    """)


@st.cache_data(show_spinner=False)
def get_ghg_submissions(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT * FROM M_GHG_SUBMISSION ORDER BY SUBMISSION_YEAR DESC, GAS
    """)


@st.cache_data(show_spinner=False)
def get_ghg_inventory(df_db_schema, fingerprint, submission, gas):
    return run_query(fingerprint, f"""
        SELECT SECTOR_CODE, SECTOR_NAME, PERIOD, KT_CO2E, IS_TOTAL_ROW
        FROM M_GHG_INVENTORY
        WHERE SUBMISSION_YEAR = '{submission}' AND GAS = '{gas}'
        ORDER BY PERIOD, SECTOR_CODE
    """)


@st.cache_data(show_spinner=False)
def get_ghg_national(df_db_schema, fingerprint, submission, gas):
    """The published national totals — gross and net as separate columns.

    Never derived by summing the sector rows: the workbook prints subtotals
    beside their own components, and summing them all inflates the figure
    ninefold.
    """
    return run_query(fingerprint, f"""
        SELECT PERIOD, KT_CO2E_NET, KT_CO2E_GROSS
        FROM M_GHG_NATIONAL
        WHERE SUBMISSION_YEAR = '{submission}' AND GAS = '{gas}'
        ORDER BY PERIOD
    """)


@st.cache_data(show_spinner=False)
def get_ghg_cross_source(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT PERIOD, PUBLISHER, BASIS, KT_CO2E
        FROM M_GHG_CROSS_SOURCE ORDER BY PERIOD, PUBLISHER
    """)


@st.cache_data(show_spinner=False)
def get_ghg_regional(df_db_schema, fingerprint, year):
    return run_query(fingerprint, f"""
        SELECT REGION, sum(KT_CO2E) AS KT_CO2E
        FROM M_GHG_REGIONAL
        WHERE PERIOD = '{year}' AND lower(REGION) <> 'total'
        GROUP BY ALL ORDER BY KT_CO2E DESC
    """)


@st.cache_data(show_spinner=False)
def get_env_accounts(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT ACCOUNT_NAME, PERIOD, sum(VALUE) AS VALUE
        FROM M_ENV_ACCOUNT WHERE VALUE IS NOT NULL
        GROUP BY ALL ORDER BY PERIOD
    """)


@st.cache_data(show_spinner=False)
def get_parcels(df_db_schema, fingerprint, intents, limit=9000):
    """Cadastral parcels with their exterior ring, ready for a PolygonLayer.

    Bounded by `limit` because pydeck draws every polygon it is handed and a
    browser will not keep up past roughly ten thousand rings. The count actually
    drawn is shown beside the map rather than left for the reader to guess.
    """
    where = f"AND PARCEL_INTENT IN ({sql_list(intents)})" if intents else ""
    return run_query(fingerprint, f"""
        SELECT PARCEL_ID, APPELLATION, PARCEL_INTENT, AREA_M2,
               LAT, LON, RING_JSON, RING_POINTS
        FROM M_LINZ_PARCEL
        WHERE RING_POINTS >= 3 {where}
        ORDER BY AREA_M2 DESC NULLS LAST
        LIMIT {limit}
    """)


@st.cache_data(show_spinner=False)
def get_parcel_intents(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT PARCEL_INTENT, count(*) AS PARCEL_COUNT, sum(AREA_M2) AS AREA_M2
        FROM M_LINZ_PARCEL GROUP BY ALL ORDER BY PARCEL_COUNT DESC
    """)


@st.cache_data(show_spinner=False)
def get_address_h3(df_db_schema, fingerprint, resolution):
    return run_query(fingerprint, f"""
        SELECT H3_CELL, ADDRESS_COUNT, SUBURB, TERRITORIAL_AUTHORITY, LAT, LON
        FROM M_LINZ_ADDRESS_H3 WHERE H3_RES = {resolution}
    """)


@st.cache_data(show_spinner=False)
def get_addresses(df_db_schema, fingerprint, text, limit=5000):
    where = ""
    if text:
        safe = text.replace("'", "''")
        where = f"WHERE FULL_ADDRESS ILIKE '%{safe}%'"
    return run_query(fingerprint, f"""
        SELECT ADDRESS_ID, FULL_ADDRESS, SUBURB, TOWN_CITY,
               TERRITORIAL_AUTHORITY, LAT, LON
        FROM M_LINZ_ADDRESS {where} ORDER BY FULL_ADDRESS LIMIT {limit}
    """)


@st.cache_data(show_spinner=False)
def get_title_summary(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT TITLE_TYPE, TITLE_STATUS, sum(TITLE_COUNT) AS TITLE_COUNT
        FROM M_LINZ_TITLE_SUMMARY GROUP BY ALL ORDER BY TITLE_COUNT DESC
    """)


@st.cache_data(show_spinner=False)
def get_river_indicators(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT INDICATOR, MEASURE_CLASS, count(*) AS SEGMENTS
        FROM M_MFE_RIVER_QUALITY GROUP BY ALL ORDER BY SEGMENTS DESC
    """)


@st.cache_data(show_spinner=False)
def get_river_quality(df_db_schema, fingerprint, measure_class, limit=30000):
    return run_query(fingerprint, f"""
        SELECT SEGMENT_ID, INDICATOR, MEASURE_CLASS, UNITS, VALUE,
               STREAM_ORDER, LAT, LON
        FROM M_MFE_RIVER_QUALITY
        WHERE MEASURE_CLASS = '{measure_class}' AND VALUE IS NOT NULL
        LIMIT {limit}
    """)


@st.cache_data(show_spinner=False)
def get_air_trends(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT POLLUTANT, SITE, AIRSHED, TREND_TYPE, LIKELIHOOD,
               SLOPE_PERCENT, PERIOD_START, PERIOD_END, LAT, LON
        FROM M_MFE_AIR_TREND WHERE SLOPE_PERCENT IS NOT NULL
        ORDER BY SLOPE_PERCENT
    """)


@st.cache_data(show_spinner=False)
def get_ade_population(df_db_schema, fingerprint, period):
    return run_query(fingerprint, f"""
        SELECT AREA_NAME, PERIOD, POPULATION
        FROM M_ADE_POPULATION
        WHERE PERIOD = '{period}' AND POPULATION IS NOT NULL
          AND AREA_NAME NOT ILIKE 'Total%'
        ORDER BY POPULATION DESC
    """)


@st.cache_data(show_spinner=False)
def get_ade_periods(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT PERIOD, count(*) AS AREAS, sum(POPULATION) AS TOTAL
        FROM M_ADE_POPULATION WHERE AREA_NAME NOT ILIKE 'Total%'
        GROUP BY ALL ORDER BY PERIOD
    """)


@st.cache_data(show_spinner=False)
def get_place_summary(df_db_schema, fingerprint):
    return run_query(fingerprint, "SELECT * FROM M_PLACE_SUMMARY")


@st.cache_data(show_spinner=False)
def get_place_h3(df_db_schema, fingerprint, resolution, feature_types, region):
    where = [f"H3_RES = {resolution}"]
    if feature_types:
        where.append(f"DOMINANT_TYPE IN ({sql_list(feature_types)})")
    if region != "All":
        where.append(f"REGION = '{region}'")
    return run_query(fingerprint, f"""
        SELECT H3_CELL, FEATURE_COUNT, DOMINANT_TYPE, REGION, LAT, LON
        FROM M_PLACE_H3 WHERE {' AND '.join(where)}
    """)


@st.cache_data(show_spinner=False)
def get_places(df_db_schema, fingerprint, feature_types, region, text):
    where = ["TRUE"]
    if feature_types:
        where.append(f"FEATURE_TYPE IN ({sql_list(feature_types)})")
    if region != "All":
        where.append(f"REGION = '{region}'")
    if text:
        safe = text.replace("'", "''")
        where.append(f"NAME ILIKE '%{safe}%'")
    return run_query(fingerprint, f"""
        SELECT NAME_ID, NAME, FEATURE_TYPE, STATUS, REGION AS LAND_DISTRICT,
               LAT, LON, MAORI_NAME_FLAG AS HAS_MAORI_NAME
        FROM M_PLACE_NAME WHERE {' AND '.join(where)}
        ORDER BY NAME LIMIT 8000
    """)


@st.cache_data(show_spinner=False)
def get_feature_types(df_db_schema, fingerprint):
    """The picker reads the dimension, not a truncated slice of the detail.

    Building it from `get_places` would offer only the types that survived that
    query's LIMIT — a selector that silently cannot offer half its options.
    """
    return run_query(fingerprint, """
        SELECT FEATURE_TYPE, FEATURE_COUNT FROM M_PLACE_FEATURE_TYPE
        ORDER BY FEATURE_COUNT DESC
    """)


@st.cache_data(show_spinner=False)
def get_place_by_region(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT REGION, sum(FEATURE_COUNT) AS FEATURE_COUNT
        FROM M_PLACE_BY_REGION GROUP BY ALL ORDER BY FEATURE_COUNT DESC
    """)


@st.cache_data(show_spinner=False)
def get_property_transfers(df_db_schema, fingerprint):
    return run_query(fingerprint, """
        SELECT PERIOD, SERIES_NAME, TRANSFERS FROM M_PROPERTY_TRANSFER
        WHERE TRANSFERS IS NOT NULL ORDER BY PERIOD
    """)


@st.cache_data(show_spinner=False)
def get_table(df_db_schema, fingerprint, table, org=None, limit=6000):
    where = f"WHERE ORG IN ({sql_list(org)})" if org else ""
    return run_query(fingerprint, f"SELECT * FROM {table} {where} LIMIT {limit}")


@st.cache_data(show_spinner=False)
def get_reference_doc(file_name):
    """Load a markdown reference document from disk, verbatim.

    Returns the text, or None when no copy is on the search path, so a
    deployment shipped without the document renders a short notice rather than
    failing. Nothing here interprets the file: it is displayed as the markdown
    it already is.
    """
    for folder in REFERENCE_DIRS:
        if not folder:
            continue
        path = Path(folder) / file_name
        if path.exists():
            with io.open(path, encoding="utf-8") as fh:
                return fh.read()
    return None


# ====================STATIC_METHODS====================
def build_styled_excel(df: pd.DataFrame, title: str) -> bytes:
    """A branded workbook: title bar, styled header, frozen panes, auto-filter."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data", startrow=2)
        wb, ws = writer.book, writer.sheets["Data"]
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=max(1, len(df.columns)))
        cell = ws.cell(row=1, column=1)
        cell.value = f"{title} — Mid-Range Organisations (Celnic Consulting)"
        cell.font = Font(bold=True, size=13, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2A78D6")
        cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 26

        thin = Side(style="thin", color="D5DBE3")
        for col_idx, name in enumerate(df.columns, start=1):
            h = ws.cell(row=3, column=col_idx)
            h.font = Font(bold=True, color="10161D")
            h.fill = PatternFill("solid", fgColor="EEF1F5")
            h.border = Border(bottom=Side(style="medium", color="2A78D6"))
            width = max(11, min(46, int(df[name].astype(str).str.len().max() or 11) + 3,
                                ))
            ws.column_dimensions[h.column_letter].width = max(width, len(str(name)) + 3)

        band = PatternFill("solid", fgColor="F7F8FA")
        for r in range(4, 4 + len(df)):
            for c in range(1, len(df.columns) + 1):
                cc = ws.cell(row=r, column=c)
                cc.border = Border(bottom=thin)
                if r % 2 == 0:
                    cc.fill = band

        ws.freeze_panes = "A4"
        if len(df.columns):
            ws.auto_filter.ref = (
                f"A3:{ws.cell(row=3, column=len(df.columns)).column_letter}"
                f"{3 + len(df)}")
    return buffer.getvalue()


def render_table_with_export(df: pd.DataFrame, title: str, key: str,
                             synthetic: bool = False) -> None:
    """A 📋 heading with the 📥 Excel button right-justified on the same row."""
    head, button = st.columns([6, 1])
    with head:
        st.markdown(f"**📋 {'🔶 ' if synthetic else ''}{title}**  \n"
                    f"<span style='color:{MUTED};font-size:0.85em'>"
                    f"{len(df):,} rows</span>", unsafe_allow_html=True)
    with button:
        if not df.empty:
            st.download_button(
                "📥 Excel", data=build_styled_excel(df, title),
                file_name=f"{key}.xlsx", key=f"dl_{key}",
                mime="application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet",
                width='stretch')
    if df.empty:
        st.info("No rows for the current filters.")
    else:
        st.dataframe(df, width='stretch', hide_index=True, height=340)


def base_layout(fig: go.Figure, title: str = "", ytitle: str = "") -> go.Figure:
    """Recessive grid and axes, one y-axis, tabular numerals."""
    fig.update_layout(
        title=dict(text=title, font=dict(size=15, color=INK)) if title else None,
        margin=dict(l=8, r=8, t=44 if title else 12, b=8),
        height=380,
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color=INK, size=12),
        legend=dict(orientation="h", yanchor="bottom", y=1.0, x=0,
                    font=dict(size=11)),
        hovermode="x unified",
    )
    fig.update_xaxes(showgrid=False, linecolor=GRID, tickfont=dict(color=MUTED))
    fig.update_yaxes(gridcolor=GRID, zeroline=False, linecolor=GRID,
                     title=ytitle, tickfont=dict(color=MUTED))
    return fig


def bar_chart(df, x, y, title, colour=REAL, ytitle="", top=15):
    """Thin horizontal bars, 4px rounded ends anchored to the baseline."""
    d = df.head(top).iloc[::-1]
    fig = go.Figure(go.Bar(
        y=d[x], x=d[y], orientation="h", marker=dict(color=colour),
        marker_cornerradius=4, hovertemplate="%{y}<br>%{x:,.0f}<extra></extra>"))
    fig = base_layout(fig, title, ytitle)
    fig.update_layout(height=max(300, 26 * len(d) + 90), bargap=0.28)
    fig.update_xaxes(showgrid=True, gridcolor=GRID)
    fig.update_yaxes(showgrid=False)
    return fig


def line_chart(series_list, title, ytitle=""):
    """2px lines, ≥8px markers, categorical hues in fixed order."""
    fig = go.Figure()
    for i, (name, x, y) in enumerate(series_list):
        fig.add_trace(go.Scatter(
            x=x, y=y, name=name, mode="lines",
            line=dict(width=2, color=SERIES_HUES[i % len(SERIES_HUES)]),
            connectgaps=False,   # suppressed periods are a GAP, never zero
            hovertemplate="%{y:,.2f}<extra>" + str(name) + "</extra>"))
    return base_layout(fig, title, ytitle)


def fmt(value, prefix="", suffix="", dp=0):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "—"
    try:
        return f"{prefix}{float(value):,.{dp}f}{suffix}"
    except (TypeError, ValueError):
        return str(value)


def render_header():
    """Hazard-striped provenance banner shown above every tab.

    Drawn as raw HTML rather than st.warning so the stripes survive both
    Streamlit themes: a 3px black frame, a 14px yellow-and-black 45-degree
    stripe band top and bottom, and a solid #FFD100 panel between them.

    The wording is the Celnic house form, shared with the MSD and Kāinga Ora
    platforms so the three read as one family. Two clauses matter beyond the
    boilerplate: the purpose, which says why the application exists at all, and
    the instruction to use the original figures rather than these — because a
    reader who takes a number off a chart here and puts it in a report has
    misunderstood what this is for.

    This platform names its seven agencies rather than pointing at the
    provenance tab, which is the one deliberate departure from the Kāinga Ora
    form. Naming them risks a stale list if another source is added; it was
    asked for, and the trade is recorded here so the next person knows it was a
    choice.

    The application is built from public releases but is not published by, nor
    endorsed by, the agencies that produced them, so that is stated before any
    figure is shown rather than buried in a footnote.
    """
    st.html(
        """
        <div style="border:3px solid #111; border-radius:6px; overflow:hidden;
                    margin:0 0 14px 0; font-family:sans-serif;">
          <div style="height:14px; background:repeating-linear-gradient(
                        45deg, #FFD100 0 14px, #111 14px 28px);"></div>
          <div style="background:#FFD100; color:#111; padding:12px 16px;">
            <div style="font-weight:800; font-size:15px; letter-spacing:.02em;">
              &#9888;&#65039; BUILT FROM NEW ZEALAND GOVERNMENT DATA &mdash;
              NOT AN OFFICIAL GOVERNMENT PRODUCT
            </div>
            <div style="font-size:13.5px; line-height:1.5; margin-top:6px;">
              Figures are reproduced from public releases by the
              <b>Department of Internal Affairs</b>, the
              <b>New Zealand Customs Service</b>, <b>Stats NZ</b>, the
              <b>Ministry of Foreign Affairs and Trade</b>,
              <b>Land Information New Zealand</b>, the
              <b>Ministry for the Environment</b> and the
              <b>Ministry of Health</b>.
              This application is produced independently by
              <b>Celnic Consulting</b> for the purpose of showing the benefits of
              Flipping the Data Team, and
              <b>does not represent the views, policy or official statistics of
              those departments</b>.
              <b>Please refer to the original figures in the data source and do
              not rely on these.</b>
              Every original source file, with its download date and checksum,
              is listed in the <b>&#128270; Data &amp; Provenance</b> tab.
            </div>
          </div>
          <div style="height:14px; background:repeating-linear-gradient(
                        45deg, #FFD100 0 14px, #111 14px 28px);"></div>
        </div>
        """
    )


def render_attribution():
    """CC BY attribution, shown once at the foot of the page.

    This does not replace the hazard-striped banner above: that one states what
    the application is and is not, and it stays at the top where a reader meets
    it before any figure. This one carries the licence obligation — attribution
    to the publishers and a statement that the data has been modified — and sits
    at the bottom, collapsed, because a licence notice is a reference the reader
    goes to rather than a warning the reader must be stopped by.

    The link points at the repository copy rather than a bundled file, so the
    full manifest travels with the deployed app on Streamlit Community Cloud,
    where only the app directory and the extract are on disk.
    """
    with st.expander("Data sources & attribution"):
        st.markdown(
            "Built on open data used under **CC BY 4.0** and **CC BY 3.0 NZ** "
            "(the Charities Register, the DIA gaming machine proceeds summary, "
            "and the New Zealand Treaties Online register) — every licence "
            "verified at source, 30 August 2026. This work is based on/includes "
            "Toitū Te Whenua Land Information New Zealand data which are "
            "licensed by Toitū Te Whenua Land Information New Zealand for "
            "re-use under the Creative Commons Attribution 4.0 International "
            "licence. Data is modified and partly synthetic; demonstration of "
            "method, not published statistics. Full provenance: "
            "[ATTRIBUTION.md](https://github.com/celnicconsulting/mid_range_org/blob/main/ATTRIBUTION.md)."
        )


# ====================SIDEBAR====================
def render_sidebar(df_db_schema, fingerprint):
    st.sidebar.title("🏛️ Mid-Range Org")
    st.sidebar.caption("Seven New Zealand agencies, one public-data platform.")

    st.sidebar.markdown("---")
    st.sidebar.subheader("Organisation")
    orgs = st.sidebar.multiselect(
        "Filter every tab by agency",
        options=list(ORG_NAMES),
        default=list(ORG_NAMES),
        format_func=lambda o: f"{o} — {ORG_NAMES[o]}")
    if not orgs:
        orgs = list(ORG_NAMES)

    st.sidebar.markdown("---")
    st.sidebar.markdown(
        f"""**PROVENANCE**

<span style='color:{REAL}'>●</span> **REAL** — a figure an agency published
<span style='color:{DERIVED}'>●</span> **DERIVED** — computed here
<span style='color:{COMPARE}'>●</span> **COMPARE** — a second agency's measure
<span style='color:{SYNTHETIC}'>🔶</span> **SYNTHETIC** — modelled, not measured
◍ **SURVEY** — carries confidence intervals

Suppressed values are drawn as a **gap**, never as zero.""",
        unsafe_allow_html=True)

    st.sidebar.markdown("---")
    summary = get_platform_summary(df_db_schema, fingerprint)
    if not summary.empty:
        vals = dict(zip(summary["METRIC"], summary["VALUE"]))
        st.sidebar.caption(
            f"{fmt(vals.get('source_files'))} source files · "
            f"{fmt(vals.get('organisations'))} agencies\n\n"
            f"{fmt(vals.get('raw_rows'))} rows landed in RAW\n\n"
            "Built by Celnic Consulting from public open data.")
    return orgs


# ====================TABS====================
def render_tab_overview(df_db_schema, fingerprint, orgs):
    """Six KPI metrics, a rows-per-agency bar, a coverage strip, the register."""
    st.subheader("Seven agencies, one public record")
    summary = get_platform_summary(df_db_schema, fingerprint)
    vals = dict(zip(summary["METRIC"], summary["VALUE"])) if not summary.empty else {}

    cols = st.columns(6)
    for col, (key, label) in zip(cols, [
            ("source_files", "Source files"), ("raw_rows", "Rows landed"),
            ("organisations", "Agencies"), ("series_staged", "Series staged"),
            ("charities_registered", "Charities registered"),
            ("treaties_in_force", "Treaties in force")]):
        col.metric(label, fmt(vals.get(key)))

    register = get_org_register(df_db_schema, fingerprint, orgs)
    if register.empty:
        st.info("No agencies selected.")
        return

    left, right = st.columns(2)
    with left:
        d = register.copy()
        d["LABEL"] = d["ORG"] + " — " + d["ORG_NAME"]
        st.plotly_chart(
            bar_chart(d, "LABEL", "STAGED_ROWS",
                      "Rows staged per agency", REAL, "rows", top=7),
            width='stretch')
    with right:
        fig = go.Figure()
        for i, row in register.iterrows():
            if not row["FIRST_PERIOD"] or not row["LAST_PERIOD"]:
                continue
            fig.add_trace(go.Scatter(
                x=[str(row["FIRST_PERIOD"])[:4], str(row["LAST_PERIOD"])[:4]],
                y=[row["ORG"], row["ORG"]], mode="lines+markers",
                line=dict(width=2, color=REAL), marker=dict(size=9, color=REAL),
                name=row["ORG"], showlegend=False,
                hovertemplate="%{y}: %{x}<extra></extra>"))
        st.plotly_chart(base_layout(fig, "Period covered, first to last", ""),
                        width='stretch')

    st.markdown("---")
    st.markdown("**Estimated resident population — Stats NZ ADE API**")
    periods = get_ade_periods(df_db_schema, fingerprint)
    if periods.empty:
        st.info("No population estimates in the extract.")
    else:
        opts = periods["PERIOD"].tolist()
        year = st.select_slider("Estimate year", options=opts, value=opts[-1],
                                key="ade_year")
        pop = get_ade_population(df_db_schema, fingerprint, year)
        left, right = st.columns([3, 2])
        with left:
            st.plotly_chart(
                bar_chart(pop, "AREA_NAME", "POPULATION",
                          f"Estimated resident population, {year}",
                          REAL, "people", top=15),
                width='stretch')
        with right:
            st.plotly_chart(
                line_chart([("New Zealand", periods["PERIOD"],
                             periods["TOTAL"])],
                           "Total across areas held", "people"),
                width='stretch')
        st.caption(
            "Pulled through the **Stats NZ Aotearoa Data Explorer API** — the "
            "route the first build could not take. It is also the platform's "
            "newest cross-source check: the same estimate reached through the "
            "API and through the bulk CSV download page agree to within 2%, "
            "and check X4 on the Data & Provenance tab shows it.")

    st.markdown("---")
    render_table_with_export(register, "Agency register", "agency_register")


def render_tab_economy(df_db_schema, fingerprint, orgs):
    """Series picker, four metrics, a levels chart, an indexed chart, detail."""
    st.subheader("💹 Economy")
    catalog = get_series_catalog(df_db_schema, fingerprint, orgs)
    if catalog.empty:
        st.info("No economic series for the selected agencies.")
        return

    catalog["LABEL"] = (catalog["SUBJECT"].fillna("") + " · "
                        + catalog["SERIES_NAME"].fillna("").str.slice(0, 90)
                        + "  [" + catalog["SERIES_REF"] + "]")
    default = catalog.sort_values("OBS_WITH_VALUE", ascending=False).head(3)

    chosen = st.multiselect(
        "Series", options=catalog["LABEL"].tolist(),
        default=default["LABEL"].tolist(), max_selections=8,
        help="Categorical hues are assigned in fixed order; a ninth series "
             "would fold into Other rather than generate a new hue.")
    refs = catalog.loc[catalog["LABEL"].isin(chosen), "SERIES_REF"].tolist()
    if not refs:
        st.info("Choose at least one series.")
        return

    # The period bounds come from the same table the facts come from, so the
    # picker can never offer a period the trimmed fact cannot answer.
    sel = catalog[catalog["SERIES_REF"].isin(refs)]
    lo, hi = str(sel["FIRST_PERIOD"].min()), str(sel["LAST_PERIOD"].max())
    period_from, period_to = st.select_slider(
        "Period range",
        options=sorted(set(pd.concat([sel["FIRST_PERIOD"], sel["LAST_PERIOD"]])
                           .astype(str).tolist() + [lo, hi])),
        value=(lo, hi))

    data = get_series(df_db_schema, fingerprint, refs, period_from, period_to)
    if data.empty:
        st.info("No observations in that period range.")
        return

    latest = (data[data["VALUE"].notna()]
              .sort_values("PERIOD").groupby("SERIES_REF").tail(1))
    cols = st.columns(min(4, max(1, len(latest))))
    for col, (_, row) in zip(cols, latest.head(4).iterrows()):
        col.metric(str(row["SERIES_NAME"])[:38] or row["SERIES_REF"],
                   fmt(row["VALUE"], dp=1),
                   help=f"{row['UNITS']} · latest period {row['PERIOD']}")

    st.plotly_chart(
        line_chart([(str(name)[:44], g["PERIOD"], g["VALUE"])
                    for name, g in data.groupby("SERIES_NAME")],
                   "Selected series over time"),
        width='stretch')

    # Indexed to the first period every selected series shares — the honest
    # answer to "these have different units", and the reason there is no second
    # y-axis anywhere in this app.
    idx_rows = []
    common = set.intersection(*[set(g["PERIOD"]) for _, g in
                                data.groupby("SERIES_REF")]) if refs else set()
    if common:
        base_period = min(common)
        for name, g in data.groupby("SERIES_NAME"):
            g = g.sort_values("PERIOD")
            base = g.loc[g["PERIOD"] == base_period, "VALUE"]
            if base.empty or not base.iloc[0]:
                continue
            idx_rows.append((str(name)[:44], g["PERIOD"],
                             g["VALUE"] / base.iloc[0] * 100))
    if idx_rows:
        st.plotly_chart(
            line_chart(idx_rows, f"Indexed — all series = 100 at {base_period}",
                       "index"),
            width='stretch')
    else:
        st.caption("The selected series share no common period, so an indexed "
                   "comparison would have no base. Narrow the selection.")

    render_table_with_export(data, "Series detail", "economy_series")


def render_tab_trade(df_db_schema, fingerprint, orgs):
    """Trade by partner and commodity, the monthly series, the treaty register."""
    st.subheader("🌏 Trade & Treaties")
    monthly = get_trade_monthly(df_db_schema, fingerprint)
    if monthly.empty:
        st.info("No trade data in the extract.")
    else:
        accounts = sorted(monthly["ACCOUNT"].dropna().unique())
        c1, c2 = st.columns([1, 3])
        account = c1.selectbox("Account", accounts,
                               index=accounts.index("Exports")
                               if "Exports" in accounts else 0)
        periods = sorted(monthly["PERIOD"].dropna().unique())
        period_from = c2.select_slider("From period", options=periods,
                                       value=periods[max(0, len(periods) - 24)])

        latest_period = periods[-1]
        latest = monthly[monthly["PERIOD"] == latest_period]
        exp = latest.loc[latest["ACCOUNT"] == "Exports", "VALUE_NZD"].sum()
        imp = latest.loc[latest["ACCOUNT"] == "Imports", "VALUE_NZD"].sum()
        m = st.columns(4)
        m[0].metric("Exports, latest period", fmt(exp, "$"), help=latest_period)
        m[1].metric("Imports, latest period", fmt(imp, "$"), help=latest_period)
        m[2].metric("Balance", fmt(exp - imp, "$"),
                    help="Exports minus imports, same month, same unit")
        m[3].metric("Periods covered", f"{len(periods):,}")

        left, right = st.columns(2)
        by_country = get_trade_by_country(df_db_schema, fingerprint,
                                          account, period_from)
        with left:
            st.plotly_chart(
                bar_chart(by_country, "COUNTRY_LABEL", "VALUE_NZD",
                          f"{account} by partner country", REAL, "NZD"),
                width='stretch')
        by_hs2 = get_trade_by_hs2(df_db_schema, fingerprint, account, period_from)
        with right:
            st.plotly_chart(
                bar_chart(by_hs2, "HS2_LABEL", "VALUE_NZD",
                          f"{account} by commodity chapter", REAL, "NZD"),
                width='stretch')

        st.plotly_chart(
            line_chart([(acct, g["PERIOD"], g["VALUE_NZD"])
                        for acct, g in monthly.groupby("ACCOUNT")],
                       "Trade value by period — exports and imports "
                       "share a unit, so they share an axis", "NZD"),
            width='stretch')
        st.caption(
            "Border-crossing counts and merchandise trade are **not joined** on "
            "this tab: their month bases differ (processing month against trade "
            "month), and reconciling them silently would invent a comparison "
            "neither agency publishes.")

    st.markdown("---")
    decades = get_treaty_decades(df_db_schema, fingerprint)
    if not decades.empty:
        fig = go.Figure()
        for i, (ttype, g) in enumerate(decades.groupby("TREATY_TYPE")):
            fig.add_trace(go.Bar(
                x=g["DECADE_CONCLUDED"], y=g["TREATY_COUNT"], name=str(ttype),
                marker=dict(color=SERIES_HUES[i % len(SERIES_HUES)],
                            line=dict(width=2, color="rgba(255,255,255,1)")),
                marker_cornerradius=4))
        fig = base_layout(fig, "Treaties concluded by decade and type", "treaties")
        fig.update_layout(barmode="stack", bargap=0.3)
        st.plotly_chart(fig, width='stretch')

    f1, f2 = st.columns(2)
    all_treaties = get_treaties(df_db_schema, fingerprint, "All", "All")
    status = f1.selectbox("NZ status", ["All"] + sorted(
        all_treaties["NZ_STATUS"].dropna().unique().tolist()))
    ttype = f2.selectbox("Treaty type", ["All"] + sorted(
        all_treaties["TREATY_TYPE"].dropna().unique().tolist()))
    render_table_with_export(
        get_treaties(df_db_schema, fingerprint, status, ttype),
        "Treaty register", "treaties")


def render_tab_civic(df_db_schema, fingerprint, orgs):
    """Gaming proceeds, the derived per-capita measure, synthetic venues, charities."""
    st.subheader("🎲 Civic & Charitable")
    st.warning(
        "🔶 **Two elements on this tab are SYNTHETIC** — the venue-level "
        "distribution and the venue detail table. Territorial-authority totals "
        "are real, and every synthetic venue rolls up to the published figure "
        "for its district **exactly**, so any one council area can be checked "
        "against DIA's own quarterly summary.")

    national = get_gmp_national(df_db_schema, fingerprint)
    charity = get_charity_summary(df_db_schema, fingerprint)
    if national.empty:
        st.info("No gaming machine proceeds in the extract.")
        return

    last = national.iloc[-1]
    m = st.columns(4)
    m[0].metric("Quarterly GMP", fmt(last["GMP_NZD"], "$"),
                help=f"Latest quarter {last['PERIOD']}")
    m[1].metric("Districts and clusters", fmt(last["TA_COUNT"]),
                help="A cluster is its own geography — DIA merges areas with "
                     "fewer than three venues into a neighbour so individual "
                     "venues cannot be identified.")
    if not charity.empty:
        m[2].metric("Charities registered", fmt(charity.iloc[0]["REGISTERED"]))
        m[3].metric("On the register in total", fmt(charity.iloc[0]["TOTAL"]),
                    help="Deregistered charities remain in the historical extract.")

    all_ta = get_gmp_ta(df_db_schema, fingerprint, None)
    top_tas = (all_ta.groupby("TA_NAME")["GMP_NZD"].sum()
               .sort_values(ascending=False).head(10).index.tolist())
    chosen_tas = st.multiselect("Territorial authorities",
                                sorted(all_ta["TA_NAME"].unique()),
                                default=top_tas[:6])
    ta_data = get_gmp_ta(df_db_schema, fingerprint, chosen_tas or top_tas)
    st.plotly_chart(
        line_chart([(name, g["PERIOD"], g["GMP_NZD"])
                    for name, g in ta_data.groupby("TA_NAME")],
                   "Gaming machine proceeds by territorial authority", "NZD"),
        width='stretch')

    st.markdown("---")
    # The picker offers only the years the derived measure can answer. Stats NZ
    # published subnational population estimates for 2018-2021 in the release
    # this build holds; offering 2008 would return an empty chart.
    pc_years = get_per_capita_years(df_db_schema, fingerprint)
    if pc_years.empty:
        st.info("No year has both gaming proceeds and a population estimate.")
        year = None
    else:
        years = pc_years["PERIOD_YEAR"].tolist()
        year = st.select_slider("Year for the per-capita measure",
                                options=years, value=years[-1])
    per_capita = (get_gmp_per_capita(df_db_schema, fingerprint, year)
                  if year else pd.DataFrame())
    if per_capita.empty:
        st.info(f"No per-capita figures for {year}.")
    else:
        st.plotly_chart(
            bar_chart(per_capita, "TA_NAME", "GMP_PER_CAPITA",
                      f"Gaming machine proceeds per resident, {year} (derived)",
                      DERIVED, "NZD per resident"),
            width='stretch')
        st.caption(
            f"**(derived)** — annual gaming machine proceeds for each territorial "
            f"authority divided by the Stats NZ estimated resident population for "
            f"that authority in {year}. **Neither agency publishes this ratio**; "
            "it is computed here from two published figures. Districts whose "
            "year is incomplete are excluded from the comparison.")

    st.markdown("---")
    periods = sorted(all_ta["PERIOD"].dropna().unique())
    syn_period = st.select_slider("Quarter for the synthetic venue view",
                                  options=periods, value=periods[-1])
    venues = get_synthetic_venues(df_db_schema, fingerprint, syn_period,
                                  chosen_tas or top_tas)
    if not venues.empty:
        fig = go.Figure()
        for vtype, g in venues.groupby("VENUE_TYPE"):
            fig.add_trace(go.Box(
                y=g["GMP_NZD"], name=str(vtype), marker=dict(color=SYNTHETIC),
                line=dict(width=2), boxpoints=False))
        st.plotly_chart(
            base_layout(fig,
                        f"🔶 SYNTHETIC — modelled venue proceeds by venue type, "
                        f"{syn_period}", "NZD per venue"),
            width='stretch')
        st.caption(
            "🔶 **Synthetic.** Venue-level proceeds are suppressed by design — "
            "DIA clusters small areas precisely so individual venues cannot be "
            "identified. These venues are modelled with a fixed seed and carry "
            "no names, only an id, a type and a district. They exist to show the "
            "*shape* of a distribution beneath the published totals, and every "
            "district sums exactly to the figure DIA published for it.")
        render_table_with_export(venues, f"Synthetic venue proceeds, {syn_period}",
                                 "synthetic_venues", synthetic=True)

    st.markdown("---")
    sectors = get_charity_by_sector(df_db_schema, fingerprint)
    if not sectors.empty:
        st.plotly_chart(
            bar_chart(sectors, "SECTOR_NAME", "CHARITY_COUNT",
                      "Charities by sector", REAL, "charities"),
            width='stretch')

    c1, c2, c3 = st.columns([2, 2, 3])
    sector_options = (["All"] + sectors["SECTOR_NAME"].tolist()
                      if not sectors.empty else ["All"])
    sector = c1.selectbox("Sector", sector_options)
    status = c2.selectbox("Status", ["All", "Registered", "Deregistered"])
    text = c3.text_input("Name contains")
    render_table_with_export(
        get_charities(df_db_schema, fingerprint, sector, status, text),
        "Charities register", "charities")
    st.caption(
        "Officer names were never downloaded and direct contact details were "
        "dropped at staging. No person is named anywhere in this platform.")


def render_tab_health(df_db_schema, fingerprint, orgs):
    """NZHS indicators with their confidence intervals; life expectancy."""
    st.subheader("🩺 Health")
    topics = get_health_topics(df_db_schema, fingerprint)
    if topics.empty:
        st.info("No health survey data in the extract.")
        return

    c1, c2, c3, c4 = st.columns([2, 3, 2, 2])
    population = c1.selectbox("Population", sorted(topics["POPULATION"].unique()))
    scoped = topics[topics["POPULATION"] == population]
    topic = c2.selectbox("Topic", sorted(scoped["TOPIC"].unique()))
    indicators = sorted(scoped.loc[scoped["TOPIC"] == topic, "INDICATOR"].unique())
    indicator = c3.selectbox(
        "Indicator", indicators,
        format_func=lambda i: str(i).replace("_", " ").capitalize(),
        help="Names are the survey's own indicator codes; the readable "
             "definitions are in the NZ Health Survey indicator reference "
             "guide, which is registered as MOH-5 but not staged.")
    age_std = c4.toggle("Age-standardised", value=False,
                        help="Crude and age-standardised estimates are two "
                             "series and are never drawn on one axis together.")

    g1, g2 = st.columns([2, 3])
    geography = g1.selectbox(
        "Geography", ["National", "Regional council"],
        help="The extract carries every national breakdown, and regional "
             "council totals plus the four total-response ethnic groups. "
             "District Health Board geographies are not offered — DHBs were "
             "abolished on 1 July 2022.")
    # The breakdown picker reads the same table the series comes from, so it can
    # never offer a split the trimmed fact cannot answer.
    available = get_health_breakdowns(df_db_schema, fingerprint, geography)
    options = available["DEMOGRAPHIC_TYPE"].dropna().tolist() or ["All"]
    breakdown = g2.selectbox("Breakdown", options, index=0)

    data = get_health_series(df_db_schema, fingerprint, indicator, population,
                             breakdown, age_std, geography)
    if data.empty:
        st.info("No estimates for that combination. Try another breakdown.")
        return

    # An indicator is published as both a Proportion and a Number. Showing one
    # of each in adjacent metrics puts "465,000" beside "69.8%" and invites the
    # reader to compare them; the design's rule is that a survey estimate is a
    # rate, never a count. One value type at a time.
    value_types = sorted(data["VALUE_TYPE"].dropna().unique())
    preferred = next((v for v in value_types if "rop" in v or "ercent" in v),
                     value_types[0] if value_types else None)
    if len(value_types) > 1:
        preferred = st.radio("Value type", value_types,
                             index=value_types.index(preferred),
                             horizontal=True)
    data = data[data["VALUE_TYPE"] == preferred] if preferred else data
    if data.empty:
        st.info("No estimates of that value type.")
        return

    latest_year = data["YEAR_TO"].max()
    latest = data[data["YEAR_TO"] == latest_year]
    m = st.columns(4)
    for col, (_, row) in zip(m, latest.head(4).iterrows()):
        col.metric(f"◍ {row['DEMOGRAPHIC_VALUE'] if geography == 'National' else row['GEOGRAPHY_VALUE']}"[:28],
                   fmt(row["VALUE"] * 100 if row["VALUE"] <= 1 else row["VALUE"],
                       suffix="%" if row["VALUE"] <= 1 else "", dp=1),
                   help=f"{latest_year} · 95% CI {row['LOW']:.3f}–{row['HIGH']:.3f}")

    # A survey estimate drawn as a bare line asserts a precision the survey does
    # not have, so the CI band is not optional.
    fig = go.Figure()
    split_col = ("GEOGRAPHY_VALUE" if geography != "National"
                 else "DEMOGRAPHIC_VALUE")
    for i, (name, g) in enumerate(data.groupby(split_col)):
        g = g.sort_values("YEAR_TO")
        hue = SERIES_HUES[i % len(SERIES_HUES)]
        if g["LOW"].notna().any():
            fig.add_trace(go.Scatter(
                x=list(g["YEAR_TO"]) + list(g["YEAR_TO"])[::-1],
                y=list(g["HIGH"]) + list(g["LOW"])[::-1],
                fill="toself", fillcolor=SURVEY_BAND, line=dict(width=0),
                hoverinfo="skip", showlegend=False))
        fig.add_trace(go.Scatter(
            x=g["YEAR_TO"], y=g["VALUE"], name=f"◍ {name}", mode="lines+markers",
            line=dict(width=2, color=hue), marker=dict(size=8, color=hue),
            connectgaps=False))
    st.plotly_chart(
        base_layout(fig, f"◍ {indicator} — {population.lower()}s, with 95% "
                         f"confidence intervals", str(data['VALUE_TYPE'].iloc[0])),
        width='stretch')
    st.caption(
        "◍ **Survey estimates.** The shaded band is the 95% confidence interval "
        "as published. These are rates, not counts, and ethnicity is "
        "total-response throughout.")

    if not latest.empty and latest[split_col].nunique() > 1:
        d = latest.sort_values("VALUE", ascending=False)
        fig2 = go.Figure(go.Bar(
            y=d[split_col], x=d["VALUE"], orientation="h",
            marker=dict(color=REAL), marker_cornerradius=4,
            error_x=dict(type="data", symmetric=False,
                         array=(d["HIGH"] - d["VALUE"]),
                         arrayminus=(d["VALUE"] - d["LOW"]),
                         color=MUTED, thickness=1.5)))
        fig2 = base_layout(fig2, f"◍ {indicator}, {latest_year}", "")
        fig2.update_layout(height=max(300, 28 * len(d) + 90), bargap=0.3)
        st.plotly_chart(fig2, width='stretch')

    st.markdown("---")
    le = get_life_expectancy(df_db_schema, fingerprint)
    if not le.empty:
        latest_le = le[le["PERIOD"] == le["PERIOD"].max()]
        by_area = (latest_le.groupby("AREA")["LIFE_EXPECTANCY"].mean()
                   .reset_index().sort_values("LIFE_EXPECTANCY", ascending=False))
        st.plotly_chart(
            bar_chart(by_area, "AREA", "LIFE_EXPECTANCY",
                      f"Life expectancy at birth, {latest_le['PERIOD'].iloc[0]}",
                      REAL, "years", top=20),
            width='stretch')

    render_table_with_export(data, f"{indicator} detail", "health_indicator")


def render_freshwater(df_db_schema, fingerprint):
    """Real MfE river water quality by segment, plus the air-quality trends.

    The other element `design/bridge_gaps.md` marked CUT. Every point is a
    modelled value MfE published for a real river segment — modelled by the
    ministry, not by this pipeline, which is a distinction the caption makes
    rather than leaving to the colour key.
    """
    ind = get_river_indicators(df_db_schema, fingerprint)
    if ind.empty:
        st.info("No river water quality layers in the extract.")
        return

    st.markdown("**🌊 Freshwater — river water quality**")
    c1, c2 = st.columns([3, 2])
    measure = c1.selectbox(
        "Measure", ind["MEASURE_CLASS"].tolist(),
        format_func=lambda m: f"{m}  ({int(ind.loc[ind['MEASURE_CLASS'] == m, 'SEGMENTS'].iloc[0]):,} segments)")
    scale = c2.radio("Colour scale", ["Value", "Stream order"],
                     horizontal=True)

    seg = get_river_quality(df_db_schema, fingerprint, measure)
    if seg.empty:
        st.info("No segments for that measure.")
        return

    col = "VALUE" if scale == "Value" else "STREAM_ORDER"
    series = pd.to_numeric(seg[col], errors="coerce")
    lo, hi = series.quantile(0.05), series.quantile(0.95)
    span = (hi - lo) or 1.0
    d = seg.copy()
    # Degraded is worse, so this ramp runs light to dark on a single hue rather
    # than red-to-green, which would assert a good/bad judgement the state
    # classes make categorically and this numeric value does not.
    d["FILL"] = series.apply(
        lambda v: [int(232 - 190 * min(max((v - lo) / span, 0), 1)),
                   int(238 - 150 * min(max((v - lo) / span, 0), 1)),
                   int(252 - 60 * min(max((v - lo) / span, 0), 1)), 190]
        if pd.notna(v) else [180, 190, 200, 90])

    st.pydeck_chart(pdk.Deck(
        layers=[pdk.Layer(
            "ScatterplotLayer", d, get_position=["LON", "LAT"],
            get_fill_color="FILL", get_radius=90, radius_min_pixels=1,
            radius_max_pixels=6, pickable=True, stroked=False)],
        initial_view_state=pdk.ViewState(
            latitude=-41.10, longitude=175.20, zoom=7.4, pitch=0),
        tooltip={"html": "<b>{MEASURE_CLASS}</b><br/>{VALUE} {UNITS}<br/>"
                         "stream order {STREAM_ORDER}<br/>segment {SEGMENT_ID}",
                 "style": {"backgroundColor": "#10161d", "color": "white"}},
        map_style="https://basemaps.cartocdn.com/gl/voyager-gl-style/style.json"))
    st.caption(
        f"**Real.** {len(seg):,} river segments, {measure}, modelled by the "
        "Ministry for the Environment for 2016–2020 and published on the MfE "
        "Data Service. The modelling is the ministry's, not this pipeline's — "
        "these are published figures, not derived ones. Clipped to the "
        "Wellington region at download, the same extent as the cadastre.")

    trends = get_air_trends(df_db_schema, fingerprint)
    if not trends.empty:
        left, right = st.columns(2)
        with left:
            worst = trends.sort_values("SLOPE_PERCENT").head(15).copy()
            worst["LABEL"] = worst["SITE"].astype(str).str.slice(0, 24) +                 " · " + worst["POLLUTANT"].astype(str)
            st.plotly_chart(
                bar_chart(worst, "LABEL", "SLOPE_PERCENT",
                          "Air quality trend by site (% a year)",
                          COMPARE, "% per year", top=15),
                width='stretch')
        with right:
            by_poll = (trends.groupby("POLLUTANT")["SLOPE_PERCENT"]
                       .mean().reset_index()
                       .sort_values("SLOPE_PERCENT"))
            st.plotly_chart(
                bar_chart(by_poll, "POLLUTANT", "SLOPE_PERCENT",
                          "Mean trend by pollutant", COMPARE, "% per year",
                          top=10),
                width='stretch')
        st.caption(
            "Negative is improving — concentrations falling. Trends are MfE's "
            "own published slopes, with their own likelihood classification; "
            "no trend is fitted here.")

    render_table_with_export(seg, f"River segments — {measure}",
                             "mfe_river_quality")


def render_tab_environment(df_db_schema, fingerprint, orgs):
    """The GHG inventory by submission and gas, the cross-source pair, accounts."""
    st.subheader("🌡️ Environment")
    submissions = get_ghg_submissions(df_db_schema, fingerprint)
    if submissions.empty:
        st.info("No greenhouse gas inventory in the extract.")
        return

    c1, c2 = st.columns(2)
    submission = c1.selectbox(
        "Inventory submission", sorted(submissions["SUBMISSION_YEAR"].unique(),
                                       reverse=True),
        help="One submission at a time. Each edition restates the whole "
             "1990-onward series on its own methodology and GWP basis, so "
             "splicing two makes a trend out of a methodology change.")
    gases = sorted(submissions.loc[
        submissions["SUBMISSION_YEAR"] == submission, "GAS"].unique())
    gas = c2.selectbox("Gas", gases,
                       index=next((i for i, g in enumerate(gases)
                                   if "all" in str(g).lower()), 0))

    basis = st.radio(
        "Accounting basis", ["Gross", "Net (after LULUCF)"], horizontal=True,
        help="Gross excludes land use, land-use change and forestry; net "
             "includes it. New Zealand's forestry sector is a large sink, so "
             "the two differ by roughly a quarter. They are two measures and "
             "are never drawn on one axis together.")
    basis_col = "KT_CO2E_GROSS" if basis == "Gross" else "KT_CO2E_NET"

    inv = get_ghg_inventory(df_db_schema, fingerprint, submission, gas)
    by_year = get_ghg_national(df_db_schema, fingerprint, submission, gas)
    if inv.empty:
        st.info("No rows for that submission and gas.")
        return

    m = st.columns(4)
    if not by_year.empty:
        latest = by_year.iloc[-1]
        base = by_year.iloc[0]
        m[0].metric(f"{basis} emissions, {latest['PERIOD']}",
                    fmt(latest[basis_col], suffix=" kt"))
        m[1].metric(f"{basis} emissions, {base['PERIOD']}",
                    fmt(base[basis_col], suffix=" kt"))
        change = ((latest[basis_col] - base[basis_col]) / base[basis_col] * 100
                  if base[basis_col] else None)
        m[2].metric(f"Change since {base['PERIOD']}", fmt(change, suffix="%", dp=1))
        m[3].metric("Submission", str(submission))

    sectors = inv[~inv["IS_TOTAL_ROW"]].copy()
    # Top-level sectors only — the workbook nests subtotals beneath their own
    # components, and stacking every level would draw the same emissions
    # several times.
    sectors = sectors[sectors["SECTOR_CODE"].astype(str).str.match(r"^\d+\.?\s*[A-Za-z]*$")]
    if not sectors.empty:
        fig = go.Figure()
        for i, (name, g) in enumerate(sectors.groupby("SECTOR_NAME")):
            g = g.sort_values("PERIOD")
            fig.add_trace(go.Scatter(
                x=g["PERIOD"], y=g["KT_CO2E"], name=str(name)[:40],
                mode="lines", stackgroup="one",
                line=dict(width=2, color=SERIES_HUES[i % len(SERIES_HUES)])))
        st.plotly_chart(
            base_layout(fig, f"Emissions by sector, {gas}, "
                             f"{submission} submission", "kt CO₂-e"),
            width='stretch')

    cross = get_ghg_cross_source(df_db_schema, fingerprint)
    if not cross.empty:
        fig = go.Figure()
        for i, (publisher, g) in enumerate(cross.groupby("PUBLISHER")):
            g = g.sort_values("PERIOD")
            fig.add_trace(go.Scatter(
                x=g["PERIOD"], y=g["KT_CO2E"], name=str(publisher),
                mode="lines+markers",
                line=dict(width=2, color=[REAL, COMPARE][i % 2]),
                marker=dict(size=8)))
        st.plotly_chart(
            base_layout(fig, "Cross-source — two publishers of national "
                             "emissions", "kt CO₂-e"),
            width='stretch')
        st.caption(
            "MfE reports **territorial** emissions on the UNFCCC basis; Stats NZ "
            "reports **production** emissions by industry and household. The two "
            "are drawn as two series and never averaged — the divergence is "
            "information about the accounting bases, not an error in either.")

    left, right = st.columns(2)
    regional = get_ghg_regional(df_db_schema, fingerprint,
                                str(by_year["PERIOD"].iloc[-1])
                                if not by_year.empty else "2020")
    with left:
        if not regional.empty:
            st.plotly_chart(
                bar_chart(regional, "REGION", "KT_CO2E",
                          "Emissions by region (Stats NZ)", COMPARE, "kt CO₂-e"),
                width='stretch')
    accounts = get_env_accounts(df_db_schema, fingerprint)
    with right:
        if not accounts.empty:
            st.plotly_chart(
                line_chart([(str(a)[:34].replace("_", " ").title(),
                             g["PERIOD"], g["VALUE"])
                            for a, g in accounts.groupby("ACCOUNT_NAME")],
                           "Environmental-economic accounts"),
                width='stretch')

    st.markdown("---")
    render_freshwater(df_db_schema, fingerprint)

    st.markdown("---")
    render_table_with_export(inv, f"Inventory detail — {gas}, {submission}",
                             "ghg_inventory")


def render_cadastre(df_db_schema, fingerprint):
    """The LINZ cadastre: real parcel polygons over real address-point density.

    This is the element `design/bridge_gaps.md` marked CUT when no API key was
    available. Both layers are real LINZ data — no synthetic geography anywhere
    on this tab.

    Two extents, stated rather than implied: the polygons cover Wellington City
    inner suburbs because a browser will not draw more rings than that, and the
    address density covers the whole Wellington region because an H3 cell costs
    the same whatever it aggregates.
    """
    intents = get_parcel_intents(df_db_schema, fingerprint)
    c1, c2, c3 = st.columns([3, 2, 2])
    chosen = c1.multiselect(
        "Parcel intent", intents["PARCEL_INTENT"].dropna().tolist(),
        default=[i for i in ["Fee Simple Title", "Road"]
                 if i in set(intents["PARCEL_INTENT"].dropna())])
    resolution = c2.slider("Address H3 resolution", 7, 10, 8,
                           help="Derived from the real address points, not a "
                                "bridge table of invented centroids.")
    show_addresses = c3.toggle("Address density", value=True)

    parcels = get_parcels(df_db_schema, fingerprint, chosen)
    cells = get_address_h3(df_db_schema, fingerprint, resolution)         if show_addresses else pd.DataFrame()

    m = st.columns(4)
    m[0].metric("Parcels drawn", fmt(len(parcels)))
    m[1].metric("Parcels held", fmt(intents["PARCEL_COUNT"].sum()),
                help="Wellington City inner extent, in the published extract.")
    m[2].metric("Address cells", fmt(len(cells)) if not cells.empty else "—")
    m[3].metric("Addresses aggregated",
                fmt(cells["ADDRESS_COUNT"].sum()) if not cells.empty else "—")

    layers = []
    if not cells.empty:
        peak = float(cells["ADDRESS_COUNT"].max()) or 1.0
        d = cells.copy()
        # Single hue, light to dark — never a rainbow.
        d["FILL"] = d["ADDRESS_COUNT"].apply(
            lambda v: [int(214 - 172 * (v / peak)),
                       int(230 - 110 * (v / peak)),
                       int(250 - 40 * (v / peak)), 150])
        layers.append(pdk.Layer(
            "H3HexagonLayer", d, get_hexagon="H3_CELL", get_fill_color="FILL",
            get_line_color=[255, 255, 255, 25], pickable=True, stroked=True,
            filled=True, extruded=False, line_width_min_pixels=0))

    if not parcels.empty:
        pl = parcels.copy()
        pl["POLYGON"] = pl["RING_JSON"].apply(
            lambda r: json.loads(r) if r else [])
        pl = pl[pl["POLYGON"].map(len) >= 3]
        layers.append(pdk.Layer(
            "PolygonLayer", pl, get_polygon="POLYGON",
            get_fill_color=[42, 120, 214, 55],
            get_line_color=[42, 120, 214, 205],
            line_width_min_pixels=1, pickable=True, stroked=True, filled=True))

    if layers:
        st.pydeck_chart(pdk.Deck(
            layers=layers,
            initial_view_state=pdk.ViewState(
                latitude=-41.2865, longitude=174.7762, zoom=12.5, pitch=0),
            tooltip={"html": "<b>{APPELLATION}</b><br/>{PARCEL_INTENT}<br/>"
                             "{AREA_M2} m²<br/>{ADDRESS_COUNT} addresses "
                             "{SUBURB}",
                     "style": {"backgroundColor": "#10161d", "color": "white"}},
            map_style="https://basemaps.cartocdn.com/gl/voyager-gl-style/style.json"))
    else:
        st.info("No parcels for the selected intents.")

    st.caption(
        f"**Real LINZ cadastre.** {len(parcels):,} parcel polygons drawn of "
        f"{int(intents['PARCEL_COUNT'].sum()):,} in the extract; rings are "
        "simplified to 1 m precision and at most 24 vertices, which is finer "
        "than a screen pixel at this zoom. Address density aggregates all "
        "215,703 regional address points. Ownership names were never "
        "downloaded — this is the titles-without-owners layer, privacy by "
        "design. Sourced from Toitū Te Whenua Land Information New Zealand "
        "data. Crown copyright reserved.")

    titles = get_title_summary(df_db_schema, fingerprint)
    left, right = st.columns(2)
    with left:
        if not titles.empty:
            st.plotly_chart(
                bar_chart(titles, "TITLE_TYPE", "TITLE_COUNT",
                          "Property titles by estate type", REAL, "titles"),
                width='stretch')
    with right:
        if not intents.empty:
            st.plotly_chart(
                bar_chart(intents, "PARCEL_INTENT", "PARCEL_COUNT",
                          "Parcels by intent", REAL, "parcels"),
                width='stretch')

    text = st.text_input("Address contains", key="addr_search")
    render_table_with_export(
        get_addresses(df_db_schema, fingerprint, text),
        "Address points", "linz_addresses")


def render_tab_places(df_db_schema, fingerprint, orgs):
    """The Gazetteer map, features by region, property transfers."""
    st.subheader("🗺️ Places & Property")

    view = st.radio(
        "Map layer", ["Cadastre — parcels & address density", "Place names"],
        horizontal=True,
        help="The cadastre is the LINZ packet's showpiece and needed a Data "
             "Service key. The place-names map covers the whole country; the "
             "cadastre covers Wellington.")

    if view.startswith("Cadastre"):
        render_cadastre(df_db_schema, fingerprint)
        st.markdown("---")

    if not view.startswith("Cadastre"):
        st.caption(
            "The New Zealand Gazetteer — real, national, and the layer this map "
            "was built on before the LINZ Data Service key was available.")

    summary = get_place_summary(df_db_schema, fingerprint)
    transfers = get_property_transfers(df_db_schema, fingerprint)
    if not summary.empty:
        s = summary.iloc[0]
        m = st.columns(4)
        m[0].metric("Named features", fmt(s["TOTAL"]))
        m[1].metric("Official names", fmt(s["OFFICIAL"]))
        m[2].metric("With a Māori name", fmt(s["MAORI_NAMED"]))
        m[3].metric("Feature types", fmt(s["FEATURE_TYPES"]))

    regions = get_place_by_region(df_db_schema, fingerprint)
    c1, c2, c3 = st.columns([3, 2, 2])
    feature_types = get_feature_types(df_db_schema, fingerprint)
    types = feature_types["FEATURE_TYPE"].dropna().tolist()
    chosen_types = c1.multiselect("Feature type", types,
                                  default=[t for t in
                                           ["Stream", "Hill", "Place", "Bay"]
                                           if t in types])
    region_options = (["All"] + regions["REGION"].tolist()
                      if not regions.empty else ["All"])
    region = c2.selectbox("Land district", region_options,
                          help="The Gazetteer's own `region` column is empty "
                               "for 53,884 of its 54,722 rows, so the LINZ land "
                               "district is used as the areal unit instead.")
    resolution = c3.slider("H3 resolution", 4, 8, 6,
                           help="Lower is coarser. Cells are derived from the "
                                "Gazetteer's own coordinates.")

    cells = get_place_h3(df_db_schema, fingerprint, resolution, chosen_types,
                         region)
    if cells.empty:
        st.info("No cells for those filters.")
    else:
        peak = float(cells["FEATURE_COUNT"].max()) or 1.0
        cells = cells.copy()
        # Single hue, light to dark — never a rainbow.
        cells["FILL"] = cells["FEATURE_COUNT"].apply(
            lambda v: [
                int(210 - 168 * (v / peak)),
                int(228 - 108 * (v / peak)),
                int(250 - 36 * (v / peak)),
                190])
        layer = pdk.Layer(
            "H3HexagonLayer", cells, get_hexagon="H3_CELL",
            get_fill_color="FILL", get_line_color=[255, 255, 255, 40],
            pickable=True, stroked=True, filled=True, extruded=False,
            line_width_min_pixels=1)
        st.pydeck_chart(pdk.Deck(
            layers=[layer],
            initial_view_state=pdk.ViewState(
                latitude=-41.2865, longitude=174.7762, zoom=5, pitch=0),
            tooltip={"html": "<b>{FEATURE_COUNT}</b> named features<br/>"
                             "{DOMINANT_TYPE}<br/>{REGION}",
                     "style": {"backgroundColor": "#10161d", "color": "white"}},
            map_style="https://basemaps.cartocdn.com/gl/voyager-gl-style/style.json"))
        st.caption(
            "Sourced from Toitū Te Whenua Land Information New Zealand data. "
            "Crown copyright reserved.")

    left, right = st.columns(2)
    with left:
        if not regions.empty:
            st.plotly_chart(
                bar_chart(regions, "REGION", "FEATURE_COUNT",
                          "Named features by land district", REAL, "features"),
                width='stretch')
    with right:
        if not transfers.empty:
            st.plotly_chart(
                line_chart([(str(n)[:40], g["PERIOD"], g["TRANSFERS"])
                            for n, g in transfers.groupby("SERIES_NAME")][:4],
                           "Property transfers by quarter", "transfers"),
                width='stretch')

    text = st.text_input("Place name contains")
    render_table_with_export(
        get_places(df_db_schema, fingerprint, chosen_types, region, text),
        "Gazetteer detail", "gazetteer")


def render_tab_build_notes(df_db_schema, fingerprint, orgs):
    """Build Notes — the build write-up, rendered from its own markdown file.

      H2  Build Notes                            + caption
      3:1 header row: 📐 Design, build and validation | download the markdown
      the document, verbatim

    The text is read at run time rather than held in this module, so the tab
    always shows the current write-up and there is never a second copy of it to
    fall out of date with the build it describes.
    """
    st.header("Build Notes")
    st.caption(
        "How the platform was designed and built: the transformation layer, the "
        "mart contract, the synthetic data, and what deployment taught. This is "
        "the project's build document, rendered from markdown."
    )

    doc = get_reference_doc(BUILD_NOTES_DOC)
    if doc is None:
        st.info(
            f"{BUILD_NOTES_DOC} is not on the reference path for this deployment. "
            "Point MID_RANGE_ORG_REFERENCE_DIR at the folder holding it."
        )
        return

    hdr, dl = st.columns([3, 1])
    with hdr:
        st.markdown("#### 📐 Design, build and validation")
    with dl:
        st.download_button(
            "📥 Markdown", data=doc.encode("utf-8"), file_name=BUILD_NOTES_DOC,
            mime="text/markdown", key="dl_build_notes_md", type="primary",
            width='stretch')

    st.markdown("---")
    st.markdown(doc)


def render_tab_provenance(df_db_schema, fingerprint, orgs):
    """Validation, the source register, coverage, and what was left out."""
    st.subheader("🔎 Data & Provenance")
    st.caption(
        "This tab **shows** the reconciliation rather than asserting it. Every "
        "tolerance below is tied to a documented mechanism, never chosen to make "
        "a check go green.")

    validation = get_table(df_db_schema, fingerprint, "M_VALIDATION_RESULTS")
    if not validation.empty:
        passed = int(validation["PASSED"].sum())
        st.metric("Reconciliation checks passed",
                  f"{passed} / {len(validation)}")
        show = validation[["CHECK_ID", "CHECK_NAME", "SCOPE", "EXPECTED",
                           "OBSERVED", "DIFFERENCE", "TOLERANCE",
                           "TOLERANCE_CAUSE", "PASSED"]]
        st.dataframe(show, width='stretch', hide_index=True, height=300)

    st.markdown("---")
    tables = {
        "Source register": ("M_PIPELINE_SOURCE", "source_register"),
        "Download manifest": ("M_PIPELINE_DOWNLOAD", "download_manifest"),
        "Coverage and gaps": ("M_PIPELINE_GAPS", "coverage_gaps"),
        "Family aliases": ("M_PIPELINE_ALIASES", "family_aliases"),
        "RAW table catalog": ("M_PIPELINE_CATALOG", "raw_catalog"),
        "Skipped RAW tables": ("M_PIPELINE_SKIPPED", "skipped_tables"),
        "Packet URL audit": ("M_PACKET_URL_AUDIT", "packet_url_audit"),
        "Excluded sources": ("M_EXCLUDED_SOURCES", "excluded_sources"),
        "Extract trim notes": ("M_EXTRACT_NOTES", "extract_notes"),
    }
    choice = st.selectbox("Register", list(tables))
    table, key = tables[choice]
    df = get_table(df_db_schema, fingerprint, table)
    if choice == "Packet URL audit":
        st.info(
            "**The research finding.** Every URL the seven work packets named "
            "was probed before any pipeline code was written. This table records "
            "which were dead and what replaced each one.")
    if choice == "Excluded sources":
        st.info(
            "Three sources named in the packets need a free registration key and "
            "were excluded from this build by the user's instruction. They are "
            "listed so a reader knows what is missing and why, rather than "
            "assuming it was never there.")
    render_table_with_export(df, choice, key)


# ====================MAIN====================
def main():
    fingerprint = _extract_fingerprint()
    if fingerprint is None:
        st.error(
            "The data extract was not found. Expected "
            "`data/mid_range_org_public.duckdb` beside the app, or "
            "`public/mid_range_org_public.duckdb` in the working project.")
        st.stop()

    df_db_schema = "MAIN"
    orgs = render_sidebar(df_db_schema, fingerprint)

    st.title("Mid-Range Organisations — New Zealand Public Data")
    render_header()

    tabs = st.tabs([
        "🏛️ Overview", "💹 Economy", "🌏 Trade & Treaties",
        "🎲 Civic & Charitable", "🩺 Health", "🌡️ Environment",
        "🗺️ Places & Property", "🔎 Data & Provenance", "🏗️ Build Notes"])

    renderers = [
        render_tab_overview, render_tab_economy, render_tab_trade,
        render_tab_civic, render_tab_health, render_tab_environment,
        render_tab_places, render_tab_provenance, render_tab_build_notes]

    for tab, render in zip(tabs, renderers):
        with tab:
            try:
                render(df_db_schema, fingerprint, orgs)
            except Exception as exc:  # noqa: BLE001
                st.error(f"This tab could not render: {exc}")

    render_attribution()


if __name__ == "__main__":
    main()
